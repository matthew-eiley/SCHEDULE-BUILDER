import pandas as pd
from random import randint, shuffle
import openpyxl
from datetime import date

EXCEL_PATH = "chief_of_staff.xlsx"
STAFF_SHEET_NAME = "STAFF"
SCHEDULE_SHEET_NAME = "SCHEDULE TEMPLATE"
TODAY = date.today()
BEER_DELIVERY_CELLS = [
    'J9', 'J10', 'J11', 'J12', 'J13'
]
SET_UP_CELLS = [
    'C19', 'C20', 'C21',
    'F19', 'F20',
    'I19', 'I20',
    'L19', 'L20',
    'O19', 'O20'
]
CELLS_5A6 =[
    'D27', 'D28', 'D29',
    'F27', 'F28',
    'H27', 'H28',
    'J27', 'J28',
    'L27', 'L28',
    'N27', 'N28',
    'P27'
]
CELLS_6A7 =[
    'D30', 'D31', 'D32',
    'F30', 'F31',
    'H30', 'H31',
    'J30', 'J31',
    'L30', 'L31',
    'N30', 'N31',
    'P30'
]
CELLS_7A8 =[
    'D33', 'D34', 'D35',
    'F33', 'F34',
    'H33', 'H34',
    'J33', 'J34',
    'L33', 'L34',
    'N33', 'N34',
    'P33'
]
CLEAN_UP_CELLS =[
    'C41', 'C42', 'C43', 'C44',
    'F41', 'F42',
    'I41', 'I42',
    'L41', 'L42',
    'O41', 'O42'
]

def read_staff_excel(file_path):
    df = pd.read_excel(io=file_path, sheet_name=STAFF_SHEET_NAME)
    return df

def add_beer_delivery(df):
    inds_used = []
    df['BEER_DELIVERY'] = False
    while len(inds_used) < 5:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, 'BEER_DELIVERY'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df

def add_set_up(df):
    inds_used = []
    df['SET_UP'] = False
    while len(inds_used) < 11:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, 'SET_UP'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df

def add_5a6(df):
    inds_used = []
    df['5A6'] = False
    while len(inds_used) < 14:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, '5A6'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df

def add_6a7(df):
    inds_used = []
    df['6A7'] = False
    while len(inds_used) < 14:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, '6A7'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df

def add_7a8(df):
    inds_used = []
    df['7A8'] = False
    while len(inds_used) < 14:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, '7A8'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df

def add_clean_up(df):
    inds_used = []
    df['CLEAN_UP'] = False
    while len(inds_used) < 12:
        rand_index = randint(0, len(df)-1)
        if (rand_index in inds_used) or (df.loc[rand_index, 'NUM_SHIFTS'] >= 2):
            continue
        inds_used.append(rand_index)
        df.loc[rand_index, 'CLEAN_UP'] = True
        df.loc[rand_index, 'NUM_SHIFTS'] += 1
    return df
    
def shift_staff(df):
    df['NUM_SHIFTS'] = 0
    df = add_beer_delivery(df)
    df = add_set_up(df)
    df = add_5a6(df)
    df = add_6a7(df)
    df = add_7a8(df)
    df = add_clean_up(df)
    return df

def get_and_shuffle(df, col_name):
    lst = df.loc[df[col_name], 'NAME'].tolist()
    shuffle(lst)
    return lst

def fill_template(df):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    source = wb[SCHEDULE_SHEET_NAME]
    ws = wb.copy_worksheet(source)
    ws.title = f"SCHEDULE {TODAY}"

    ws['B3'] = f"MUS 4À7 SCHEDULE ({TODAY})"

    beer_delivery_names = get_and_shuffle(df, 'BEER_DELIVERY')
    for i, cell in enumerate(BEER_DELIVERY_CELLS):
        ws[cell] = beer_delivery_names[i]

    set_up_names = get_and_shuffle(df, 'SET_UP')
    for i, cell in enumerate(SET_UP_CELLS):
        ws[cell] = set_up_names[i]

    names_5a6 = get_and_shuffle(df, '5A6')
    for i, cell in enumerate(CELLS_5A6):
        ws[cell] = names_5a6[i]

    names_6a7 = get_and_shuffle(df, '6A7')
    for i, cell in enumerate(CELLS_6A7):
        ws[cell] = names_6a7[i]

    names_7a8 = get_and_shuffle(df, '7A8')
    for i, cell in enumerate(CELLS_7A8):
        ws[cell] = names_7a8[i]

    clean_up_names = get_and_shuffle(df, 'CLEAN_UP')
    for i, cell in enumerate(CLEAN_UP_CELLS):
        ws[cell] = clean_up_names[i]

    not_shifted = df[df['NUM_SHIFTS'] == 0]['NAME'].tolist()
    single_shifted = df[df['NUM_SHIFTS'] == 1]['NAME'].tolist()
    double_shifted = df[df['NUM_SHIFTS'] == 2]['NAME'].tolist()

    for i, name in enumerate(not_shifted):
        ws[f'S{11+i}'] = name
    for i, name in enumerate(single_shifted):
        ws[f'U{11+i}'] = name
    for i, name in enumerate(double_shifted):
        ws[f'W{11+i}'] = name

    wb.save(EXCEL_PATH)

def main():
    df = read_staff_excel(EXCEL_PATH)
    df = shift_staff(df)
    fill_template(df)
    print(df)

if __name__ == "__main__":
    main()