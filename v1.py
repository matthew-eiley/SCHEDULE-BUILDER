import pandas as pd
from random import randint, shuffle
import openpyxl
from datetime import date

EXCEL_PATH = "chief_of_staff.xlsx"
STAFF_SHEET_NAME = "STAFF"
SCHEDULE_SHEET_NAME = "SCHEDULE TEMPLATE"
TODAY = date.today()
BEER_DELIVERY_CELLS = [
    'J9', 'J10', 'J11', 'J12', 'J13'
]
SET_UP_CELLS = [
    'C19', 'C20', 'C21',
    'F19', 'F20',
    'I19', 'I20',
    'L19', 'L20',
    'O19', 'O20'
]
CLEAN_UP_CELLS =[
    'C41', 'C42', 'C43', 'C44',
    'F41', 'F42',
    'I41', 'I42',
    'L41', 'L42',
    'O41', 'O42'
]
CELLS_5A6 =[
    'D27', 'D28', 'D29',
    'F27', 'F28',
    'H27', 'H28',
    'J27', 'J28',
    'L27', 'L28',
    'N27', 'N28',
    'P27'
]
CELLS_7A8 =[
    'D33', 'D34', 'D35',
    'F33', 'F34',
    'H33', 'H34',
    'J33', 'J34',
    'L33', 'L34',
    'N33', 'N34',
    'P33'
]
CELLS_6A7 =[
    'D30', 'D31', 'D32',
    'F30', 'F31',
    'H30', 'H31',
    'J30', 'J31',
    'L30', 'L31',
    'N30', 'N31',
    'P30'
]

def read_staff_excel(file_path):
    df = pd.read_excel(io=file_path, sheet_name=STAFF_SHEET_NAME)
    df['STATUS'] = df['STATUS'].map({'Available': True, 'Unavailable': False})
    return df

def do_beer_delivery(df):
    num_needed = 5
    num_staff = len(df)
    possible_inds = [i for i in range(num_staff)]
    inds_chosen = []

    for _ in range(num_needed):
        rand_ind = randint(0, len(possible_inds) - 1)
        inds_chosen.append(possible_inds[rand_ind])
        possible_inds.pop(rand_ind)

    df['BEER_DELIVERY'] = False
    df.loc[inds_chosen, 'BEER_DELIVERY'] = True

    return df

def do_set_and_clean_up(df):
    set_up_num_needed = 11
    clean_up_num_needed = 12
    num_staff = len(df)
    possible_inds = [i for i in range(num_staff) if df.loc[i, 'STATUS']]
    set_up_inds_chosen = []
    clean_up_inds_chosen = []

    for _ in range(set_up_num_needed):
        rand_ind = randint(0, len(possible_inds) - 1)
        set_up_inds_chosen.append(possible_inds[rand_ind])
        possible_inds.pop(rand_ind)

    for _ in range(clean_up_num_needed):
        rand_ind = randint(0, len(possible_inds) - 1)
        clean_up_inds_chosen.append(possible_inds[rand_ind])
        possible_inds.pop(rand_ind)

    df['SET_UP'] = False
    df.loc[set_up_inds_chosen, 'SET_UP'] = True

    df['CLEAN_UP'] = False
    df.loc[clean_up_inds_chosen, 'CLEAN_UP'] = True

    return df

def do_56_and_78(df):
    num_needed = 14
    num_spots_56 = num_needed - df['SET_UP'].sum()
    num_spots_78 = num_needed - df['CLEAN_UP'].sum()
    possible_inds = [i for i in range(len(df)) if (df.loc[i, 'STATUS'] and not df.loc[i, 'SET_UP'] and not df.loc[i, 'CLEAN_UP'])]
    inds_chosen_56 = []
    inds_chosen_78 = []

    for _ in range(num_spots_56):
        rand_ind = randint(0, len(possible_inds) - 1)
        inds_chosen_56.append(possible_inds[rand_ind])
        possible_inds.pop(rand_ind)

    for _ in range(num_spots_78):
        rand_ind = randint(0, len(possible_inds) - 1)
        inds_chosen_78.append(possible_inds[rand_ind])
        possible_inds.pop(rand_ind)

    df['5-6'] = False
    df.loc[inds_chosen_56, '5-6'] = True
    df.loc[df['SET_UP'], '5-6'] = True

    df['7-8'] = False
    df.loc[inds_chosen_78, '7-8'] = True
    df.loc[df['CLEAN_UP'], '7-8'] = True

    return df

def do_67(df):
    num_needed = 14
    inds_not_yet_shifted = [i for i in range(len(df)) if (df.loc[i, 'STATUS'] and not df.loc[i, '5-6'] and not df.loc[i, '7-8'])]
    num_not_yet_shifted = len(inds_not_yet_shifted)
    if num_not_yet_shifted < num_needed:
        try:
            inds_avail_for_double = [i for i in range(len(df)) if (df.loc[i, 'STATUS'] and not df.loc[i, 'BEER_DELIVERY'] and (df.loc[i, '5-6'] or df.loc[i, '7-8']) and not (df.loc[i, 'SET_UP'] or df.loc[i, 'CLEAN_UP']))]
        except: # in case not enough staff available (shouldn't happen)
            inds_avail_for_double = [i for i in range(len(df)) if (df.loc[i, 'STATUS'] and not df.loc[i, 'BEER_DELIVERY'] and (df.loc[i, '5-6'] or df.loc[i, '7-8']))]
        shuffle(inds_avail_for_double)
        doubles_needed = num_needed - num_not_yet_shifted            
        inds_chosen_for_double = inds_avail_for_double[:doubles_needed]
        inds_chosen_67 = inds_not_yet_shifted + inds_chosen_for_double
    else:
        shuffle(inds_not_yet_shifted)
        inds_chosen_67 = inds_not_yet_shifted[:num_needed]

    df['6-7'] = False
    df.loc[inds_chosen_67, '6-7'] = True

    return df

def do_num_shifts(df):
    df['NUM_SHIFTS'] = df[['5-6', '6-7', '7-8']].sum(axis=1)
    df.sort_values(by='NUM_SHIFTS', inplace=True)
    return df

def get_and_shuffle(df, col_name):
    lst = df.loc[df[col_name], 'NAME'].tolist()
    shuffle(lst)
    return lst

def fill_template(df):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    source = wb[SCHEDULE_SHEET_NAME]
    ws = wb.copy_worksheet(source)
    wb.save(EXCEL_PATH)
    ws.title = f"SCHEDULE {TODAY}"
    ws['B3'] = f"MUS 4À7 SCHEDULE ({TODAY})"

    beer_delivery_names = get_and_shuffle(df, 'BEER_DELIVERY')
    for i, cell in enumerate(BEER_DELIVERY_CELLS):
        ws[cell] = beer_delivery_names[i]
    set_up_names = get_and_shuffle(df, 'SET_UP')
    for i, cell in enumerate(SET_UP_CELLS):
        ws[cell] = set_up_names[i]
    names_5a6 = get_and_shuffle(df, '5-6')
    for i, cell in enumerate(CELLS_5A6):
        ws[cell] = names_5a6[i]
    names_6a7 = get_and_shuffle(df, '6-7')
    for i, cell in enumerate(CELLS_6A7):
        ws[cell] = names_6a7[i]
    names_7a8 = get_and_shuffle(df, '7-8')
    for i, cell in enumerate(CELLS_7A8):
        ws[cell] = names_7a8[i]
    clean_up_names = get_and_shuffle(df, 'CLEAN_UP')
    for i, cell in enumerate(CLEAN_UP_CELLS):
        ws[cell] = clean_up_names[i]

    # not_shifted = df[df['NUM_SHIFTS'] == 0]['NAME'].tolist()
    # single_shifted = df[df['NUM_SHIFTS'] == 1]['NAME'].tolist()
    # double_shifted = df[df['NUM_SHIFTS'] == 2]['NAME'].tolist()
    # for i, name in enumerate(not_shifted):
    #     ws[f'S{11+i}'] = name
    # for i, name in enumerate(single_shifted):
    #     ws[f'U{11+i}'] = name
    # for i, name in enumerate(double_shifted):
    #     ws[f'W{11+i}'] = name

    wb.save(EXCEL_PATH)

def main():
    df = read_staff_excel(EXCEL_PATH)
    df = do_beer_delivery(df)
    df = do_set_and_clean_up(df)
    df = do_56_and_78(df)
    df = do_67(df)
    df = do_num_shifts(df)
    print(df)
    fill_template(df)

if __name__ == "__main__":
    main()